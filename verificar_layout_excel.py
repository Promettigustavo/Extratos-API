import openpyxl
import time

# Aguardar um pouco para garantir que o arquivo foi fechado
time.sleep(1)

arquivo = 'exportar-Santander - Extrato 14 de novembro de 2025-2271-130137784.xlsx'

try:
    wb = openpyxl.load_workbook(arquivo)
    ws = wb.active
    
    print('=' * 80)
    print('VERIFICAÇÃO DO LAYOUT EXCEL')
    print('=' * 80)
    print(f'\nTítulo da aba: {ws.title}')
    print(f'Total de linhas: {ws.max_row}')
    print(f'Total de colunas: {ws.max_column}')
    
    print('\n' + '-' * 80)
    print('CABEÇALHO (Linha 1):')
    print('-' * 80)
    print([cell.value for cell in ws[1]])
    
    print('\n' + '-' * 80)
    print('PRIMEIRA TRANSAÇÃO (Linha 2):')
    print('-' * 80)
    print([cell.value for cell in ws[2]])
    
    print('\n' + '-' * 80)
    print('FORMATAÇÃO DO HEADER (célula A1):')
    print('-' * 80)
    cell = ws['A1']
    print(f'  Fonte: {cell.font.name} {cell.font.size}pt')
    print(f'  Negrito: {cell.font.bold}')
    print(f'  Cor de fundo: {cell.fill.start_color.rgb if cell.fill.start_color else "None"}')
    print(f'  Alinhamento: {cell.alignment.horizontal}')
    
    print('\n' + '-' * 80)
    print('LARGURAS DAS COLUNAS:')
    print('-' * 80)
    cols = ['A','B','C','D','E','F','G','H','I','J']
    for c in cols:
        print(f'  Coluna {c}: {ws.column_dimensions[c].width:.2f}')
    
    print('\n' + '=' * 80)
    print('✅ LAYOUT VERIFICADO COM SUCESSO!')
    print('=' * 80)
    
    wb.close()
    
except PermissionError:
    print('\n❌ ERRO: O arquivo está aberto no Excel.')
    print('   Por favor, feche o arquivo e execute novamente.')
except Exception as e:
    print(f'\n❌ ERRO: {e}')
    import traceback
    traceback.print_exc()
